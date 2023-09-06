from dataclasses import asdict
from openpyxl import Workbook
import pandas as pd
import wandb
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

from torch_timeseries.experiments.dlinear_experiment import DLinearExperiment
# 初始化Wandb API
api = wandb.Api()

def convert_dict(dictionary):
    converted_dict = {}
    for key, value in dictionary.items():
        converted_dict[f"config.{key}"] = value
    return converted_dict

# 查询特定配置的运行
project_name = "BiSTGNN"  # 替换为你的项目名称
config_name = "your-config-name"  # 替换为你的配置名称

# single steps report

model_types = ["DLinear", "NLinear", "MTGNN", "Crossformer", "TSMixer", "FiLM", "GRU"]
datasets = ["ExchangeRate",  "ETTm1", "ETTm2", "ETTh1", "ETTh2"] #"ExchangeRate",
horizons = [3, 6, 12, 24]

all_data = {}
runs = api.runs(path="BiSTGNN")
for run in runs:
    if run.state == "finished":
        try:
            all_data[(run.config['model_type'],run.config['dataset_type'], run.config['horizon'], 'mse')] = run.summary['mse']
            all_data[(run.config['model_type'],run.config['dataset_type'], run.config['horizon'], 'r2')] = run.summary['r2']
            all_data[(run.config['model_type'],run.config['dataset_type'], run.config['horizon'], 'r2w')] = run.summary['r2_weighted']
            all_data[(run.config['model_type'],run.config['dataset_type'], run.config['horizon'], 'mae')] = run.summary['mae']
        except:
            continue
index = pd.MultiIndex.from_tuples(all_data.keys(), names=['dataset', 'horizon', 'metric', 'model_type'])


df = pd.DataFrame(all_data.values(), index=index)
"""
(Pdb) df
                                               0
dataset horizon metric model_type               
GRU     ETTh1   12     mse         0.3308±0.0058
                       r2          0.5113±0.0118
                       r2w         0.5851±0.0072
                       mae         0.4093±0.0041
DLinear ETTh2   12     mse         0.0747±0.0007
...                                          ...
MTGNN   ETTh2   24     mae         0.2188±0.0054
                12     mse         0.0783±0.0036
                       r2          0.2924±0.0532
                       r2w         0.4056±0.0272
                       mae         0.2052±0.0065

"""

df.unstack(0).to_excel('result.xls')
"""
dataset                      Crossformer        DLinear FiLM            GRU       Informer     Informer-t          MTGNN NLinear TSMixer       TimesNet
horizon metric model_type                                                                                                                              
ETTh1   3      mae         0.3143±0.0022  0.3217±0.0013  NaN  0.3334±0.0018            NaN            NaN  0.3237±0.0059     NaN     NaN            NaN
               mse         0.2128±0.0037  0.2201±0.0012  NaN  0.2306±0.0012            NaN            NaN  0.2183±0.0075     NaN     NaN            NaN
               r2          0.6887±0.0085  0.6868±0.0009  NaN  0.6651±0.0046            NaN            NaN  0.6688±0.0118     NaN     NaN            NaN
               r2w         0.7350±0.0046  0.7259±0.0015  NaN  0.7128±0.0015            NaN            NaN  0.7280±0.0093     NaN     NaN            NaN
        6      mae         0.3678±0.0053  0.3694±0.0009  NaN  0.4058±0.0041  1.8754±0.0841  1.7567±0.1092  0.3851±0.0071     NaN     NaN  0.4048±0.0086
...                                  ...            ...  ...            ...            ...            ...            ...     ...     ...            ...
Weather 12     r2w         0.6300±0.0037            NaN  NaN            NaN            NaN            NaN  0.6216±0.0048     NaN     NaN            NaN
        24     mae         0.1768±0.0075            NaN  NaN            NaN            NaN            NaN        nan±nan     NaN     NaN            NaN
               mse         0.1382±0.0034            NaN  NaN            NaN            NaN            NaN        nan±nan     NaN     NaN            NaN
               r2          0.4352±0.0388            NaN  NaN            NaN            NaN            NaN        nan±nan     NaN     NaN            NaN
               r2w         0.5073±0.0120            NaN  NaN            NaN            NaN            NaN        nan±nan     NaN     NaN            NaN
# """

# import pdb;pdb.set_trace()
# # 创建Excel工作簿
# wb = Workbook()
# ws = wb.active

# # 将DataFrame数据转换为Excel表格的行
# for row in dataframe_to_rows(df, index=True, header=True):
#     ws.append(row)
    
# # 合并单元格
# for cell in ws['A1.expand']: 
#     cell.merge('A2')

# for cell in ws['B1':'D1']: 
#     cell.merge(cell.offset(row=1))


# for row in ws.iter_rows(min_row=1, max_row=1):
#     for cell in row: cell.alignment = Alignment(horizontal='center', vertical='center')


# ws.row_dimensions[1].height = 20 
# ws.column_dimensions['A'].width = 15
# ws.column_dimensions['B'].width = 10 
# ws.column_dimensions['C'].width = 10
# ws.column_dimensions['D'].width = 10
# wb.save('multi_indexed_cells.xlsx')



# exp = DLinearExperiment(
#         epochs=100,
#         patience=5,
#         windows=90,
#         horizon=3,
#         dataset_type="ExchangeRate",
#         device="cuda:4",
# )

# config = exp.result_related_config
# config_filter = convert_dict(config)
# runs = api.runs(path="BiSTGNN", filters=config_filter)
# print(runs[0].state== "finished")








#{'lr': 0.0003, 'device': 'cuda:4', 'epochs': 100, 'dropout': 0.1, 'horizon': 3, 'windows': 384, 'patience': 5, 'pred_len': 1, 'save_dir': './results', 'data_path': './data', 'optm_type': 'Adam', 'batch_size': 32, 'model_type': 'GRU', 'num_layers': 2, 'num_worker': 20, 'hidden_size': 128, 'scaler_type': 'StandarScaler', 'dataset_type': 'ETTh1', 'invtrans_loss': False, 'max_grad_norm': 5, 'loss_func_type': 'mse', 'l2_weight_decay': 0.0005, 'experiment_label': '1690863787'}